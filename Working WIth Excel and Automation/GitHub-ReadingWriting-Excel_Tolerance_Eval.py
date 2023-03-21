import os
import sys
import shutil
from openpyxl import load_workbook
import openpyxl
#not extendedopenpyxl for this program, the module is for saving properly 2 different fonts in a single cell
#don't really need it for my purposes in this program
#from extendedopenpyxl import load_workbook, save_workbook
import pandas as pd
import easygui
from PyQt5.QtWidgets import QApplication, QLabel, QDialog, QLineEdit, QCheckBox, QPushButton
from PyQt5.QtCore import Qt, pyqtSignal, QThread
from PyQt5.QtGui import QFont
import getpass
from openpyxl.styles import PatternFill

#Overview of Program:
#find excel file in specific locations, read them and use logic to collect certain data
#based on the data break it up in certain way in order to perform calculations
#output results based on the calculations to excel files

#main GUI class
class Actions(QDialog):

    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setGeometry(400,200,550,120)
        self.setFixedSize(self.size())
        self.setWindowFlag(Qt.WindowMinimizeButtonHint, True)

        self.setWindowTitle("Tolerance Evaluation")

        self.about = QLabel("About: Evaluate results on an FAI to see how much Tolerance they use up.", self)
        self.about.adjustSize()
        self.about.move(25, 0)

        self.instruct = QLabel("---  After program evaluates, will save results to folder on your desktop", self)
        self.instruct.setFont(QFont('Arial', 7))
        self.instruct.adjustSize()
        self.instruct.move(50, 16)

        self.part_label = QLabel("<b>Enter Part#:<B>", self)
        self.part_label.adjustSize()
        self.part_label.move(15, 35)

        self.part = QLineEdit(self)
        self.part.setFixedWidth(100)
        self.part.move(15, 50)

        self.job_label = QLabel("<b>Enter JOB#:<B>", self)
        self.job_label.adjustSize()
        self.job_label.move(145, 35)

        self.job = QLineEdit(self)
        self.job.setFixedWidth(100)
        self.job.move(145, 50)

        self.check = QCheckBox(self)
        self.check.setText("Show results overlaid to FAI?")
        self.check.move(275, 40)

        self.check1 = QCheckBox(self)
        self.check1.setText("Show compiled results separately when done?")
        self.check1.move(275, 60)

        self.button = QPushButton(self)
        self.button.setText("Start")
        self.button.clicked.connect(self.onButtonClick)
        self.button.move(15, 80)

        self.update_label = QLabel("                                                                                                  ", self)
        self.update_label.adjustSize()
        self.update_label.move(110, 85)

        self.show()

    #on start button click this activates
    def onButtonClick(self):
        try:
            #clean up the text entered into the fields in the GUI by removing excess spaces and making all capital
            a = self.part.text().strip(" ")
            a_upper = a.upper()
            b = self.job.text().strip(" ")
            b_upper = b.upper()
            #check if checkboxes are checked and put them in a variable for use later
            c = self.check.checkState()
            d = self.check1.checkState()

            #update text status in GUI for what the program is doing
            self.update_label.setText("Starting.....finding file....")

            #make sure value entered into first text field on GUI is 12 characters
            if len(a) != 12:
                msg = easygui.msgbox("Part Number or rev not correct!", "ERROR")
                self.update_label.setText("Try Again....")
            else:
                #pass info to find_excel method for checking to make sure paths exists
                excel_path, path_exists = self.find_excel(a_upper, b_upper)

                #error if path doesn't exists
                if path_exists == False:
                    msg = easygui.msgbox("Part Number not correct or Excel not saved to FAI Folder!!", "ERROR")
                    self.update_label.setText("Try Again....")

                #start worker thread if path exists
                if path_exists == True:
                    self.calc = External(excel_path, c, d)
                    self.calc.updateChanged.connect(self.onupdateChanged)
                    self.calc.exitChanged.connect(self.onexitChanged)
                    self.calc.start()
        #if something doens't work display error
        except:
            easygui.msgbox(msg="ERROR: 8008 8008", title="ERROR")

    #for finding directory paths
    def find_excel(self, a_upper, b_upper):

        #use this method to make sure paths exists for files

        filepath = "O:\\your base path"
        path_exists = True
        #if path doesn't exist, false will be returned to method that called it
        if not os.path.exists(filepath):
            path_exists = False
        else:
            #if path exists, search for excels in path (this particular code grabs the last excel
            #in the folder, as I want the last excel windows gets from the default windows alphabetical order
            test_for_xlsx = os.listdir(filepath)
            k = 0
            while k < len(test_for_xlsx):
                if ".xlsx" in test_for_xlsx[k]:
                    if "~" not in test_for_xlsx[k]:
                        partexcel = test_for_xlsx[k]
                else:
                    path_exists = False
                k += 1

        if path_exists == True:
            filepath = "O:\\your base path"

        #return path to directory and if path exists
        return(filepath, path_exists)

    #for updating text field in GUI for what program is doing
    def onupdateChanged(self, value):
        self.update_label.setText(value)

    #for terminating worker thread if an error happens, this code is not actually used but keeping on here in
    #case i want to reimplement it
    def onexitChanged(self, value1):
        if "Yes" in value1:
            self.calc.terminate()

#worker thread that reads excel and performs actions
class External(QThread):
    updateChanged = pyqtSignal(str)
    exitChanged = pyqtSignal(str)

    def __init__(self, excel_path, c, d):
        super(External, self).__init__()
        #path to excel file
        self.excel_path = excel_path
        #checkbox states from GUI if they are checked or not
        self.c = c
        self.d = d

    #worker method that auto starts when class is called (a feature of pyqt5 Qthread)
    def run(self):
        try:
            #update text in GUI
            self.updateChanged.emit("Reading FAI Excel file........")
            user = getpass.getuser()

            TEMP_EXCEL = "C:\\tolerance_evaluations"

            #if temp path doesn't exist create it, this temp path used to temporarily save excel files,
            #program deletes it later
            try:
                os.stat(TEMP_EXCEL)
            except:
                os.mkdir(TEMP_EXCEL)

            #read excel file
            self.dataframe1 = pd.ExcelFile(self.excel_path)
            self.names1 = self.dataframe1.sheet_names

            #setup some empty lists to store data, the data that gets put into these lists used to
            #find the correct row/cell that contains the data i want and then add data to correct
            #cell within that row for the specific excel files I'm using.
            item_loc = []
            ref_loc = []
            full_requirement = []
            full_result = []
            percentages = []
            tool_used = []

            #iterate through columns in sheet and gather results from the cells.  The way the data is gathered
            #from the cells and broken down is specific to my needs for this program.
            #The way I have this broken down for breaking that data up could be done better, a little too many
            #if/else statements.....but I just created the program in a hurry

            for i in self.names1:  #iterate all sheets
                #update text in GUI for program status
                self.updateChanged.emit("Reading Sheet " + str(i))
                if "FAIR 3" in i.upper():    #find all fair3 sheets
                    iterate_fair = self.dataframe1.parse(i)
                    for index, row in iterate_fair.iloc[:, 6].iteritems():    #iterate through result column in every fair 3
                        if "nan" not in str(row) and "10. Results" not in str(row):   #remove items that aren't results                    #get result
                            requirement = str(iterate_fair.iloc[index, 4])            #get requirement

                            counter = requirement.count("±") # use counter in case requirement has 2 ±'s in it

                            gdt = requirement.count("|")
                            gdt_requirement = False
                            if gdt >= 3:
                                counter = 1
                                gdt_requirement = True


                            ranged_requirement = False      #check for a ranged requirement
                            if "/" in requirement:
                                if "+" not in requirement:
                                    nominal_requirement = requirement
                                    nominal_result, ranged_tol = self.ranged_req(nominal_requirement)

                                   # print(str(nominal_result) + " " + str(ranged_tol))
                                    if "ERROR" not in str(nominal_result):
                                        counter = 1
                                        ranged_requirement = True


                            angle = False          #if requirement is an angle with no tol, then tol is 2
                            if "°" in requirement:
                                if "±" not in requirement and "/" not in requirement:
                                    if len(requirement) < 5:
                                        counter = 1
                                        angle = True

                            ranged_tolerance = False
                            if "+" in requirement and "/" in requirement and "-" in requirement:
                                counter = 1
                                ranged_tolerance = True

                            if counter == 1:              #split requirement up by nominal/tolerance for those that have ±

                                if gdt_requirement == True:
                                    tolerance_real = requirement
                                    nominal = str(0)
                                    result_raw = str(iterate_fair.iloc[index, 6])
                                    result = result_raw.strip(" ")

                                if ranged_tolerance == True:

                                    start_index = requirement.index("+")
                                    tolerance_raw = requirement[start_index + 1:].strip(" ")
                                    nominal = requirement[:start_index - 1].strip(" ")
                                    result = str(iterate_fair.iloc[index, 6])
                                    result = result.strip(" ")
                                    upper, lower = self.upper_lower_tol(tolerance_raw)

                                  #  print(nominal, result, upper, lower)

                                if ranged_requirement == True:
                                    nominal = str(nominal_result)
                                    tolerance_raw = str(ranged_tol)
                                    result = str(iterate_fair.iloc[index, 6])
                                    result = result.strip(" ")
                                else:
                                    if angle == False and ranged_requirement == False and ranged_tolerance == False and gdt_requirement == False:
                                        start_index = requirement.index("±")
                                        tolerance_raw = requirement[start_index+1:].strip(" ")
                                        nominal = requirement[:start_index].strip(" ")
                                        result = iterate_fair.iloc[index, 6]
                                        result = str(result)

                                    else:
                                        if angle == True:
                                            tolerance_raw = "2"
                                            nominal = requirement
                                            result = iterate_fair.iloc[index, 6].strip(" ")


                    #======Below block of code just for extracting and cleaning up tolerance==========================
                                if ranged_tolerance == False:

                                    if gdt_requirement == False:      #don't run this to when gdt is true otherwise will messed it up
                                        if " " in str(tolerance_raw):  #removes anything after tolerance such as "X 100°"
                                            k = 0
                                            end_whileloop = False
                                            while k < len(tolerance_raw):
                                                if " " in str(tolerance_raw[k]):
                                                    tolerance_real = tolerance_raw[0:k]
                                                    up_k = len(tolerance_raw) - k
                                                    end_whileloop = True
                                                if end_whileloop == True:
                                                    k += up_k
                                                k += 1
                                        else:
                                            tolerance_real = tolerance_raw
                                    #print(tolerance_real)

                                    k = 0          #if .010x100 no spaces, extracts tolerance at in front of x
                                    while k < len(tolerance_real):
                                        if "X" in str(tolerance_real[k]).upper():
                                            tolerance_real = tolerance_real[0:k]
                                            k += len(tolerance_real)
                                        k += 1
                                    #print(tolerance_real)

                                    k = 0                                         #removes any symbols, such as °, except "." from tolerance
                                    iterate_tol = len(tolerance_real)
                                    while k < iterate_tol:
                                        if "." not in str(tolerance_real[k]):
                                            a = tolerance_real[k]
                                            b = a.isdigit()
                                            if b != True:
                                                counter_letter = tolerance_real.count(a)
                                                tolerance_real = tolerance_real.replace(a, '')
                                                while counter_letter > 0:
                                                    iterate_tol -= 1
                                                    counter_letter -= 1
                                                k -= 1
                                        k += 1

                                    #print(tolerance_real)
                    # ======Above block of code just for extracting and cleaning up tolerance===============================



                    #===========Below block of code cleans up NOMINAL==========================================

                              #  debug = str(iterate_fair.iloc[index, 0])
                               # if "80" in str(debug):
                               #     print()

                                if len(nominal) > 3:
                                    if "X" in str(nominal[1].upper()):     #removes things like "6X" from front of dim
                                        nominal = nominal[2:].strip(" ")
                                    if "X" in str(nominal[2].upper()):     #removes things like "16X" from front of dim
                                        nominal = nominal[3:].strip(" ")
                                #print(nominal)

                                if gdt_requirement == False:  #don't run space code when gdt nominal is true
                                    total_nom = len(nominal)       #when space detected, make that nominal, so removes things like "R" if there's a space inbetween
                                    while total_nom > 0:
                                        if " " in str(nominal[total_nom-1]):
                                            nominal = nominal[total_nom-1:len(nominal)].strip(" ")
                                            total_nom -= total_nom
                                        else:
                                            total_nom -= 1

                                    #print(nominal)



                                k = 0  # removes any symbols, such as ° or "R", except "." from nominal if no space inbetween
                                iterate_nom = len(nominal)
                                while k < iterate_nom:
                                    if "." not in str(nominal[k]):
                                        a = nominal[k]
                                        b = a.isdigit()
                                        if b != True:
                                            counter_letter = nominal.count(a)
                                            nominal = nominal.replace(a, '')
                                            while counter_letter > 0:
                                                iterate_nom -= 1
                                                counter_letter -= 1
                                            k -= 1
                                    k += 1
                                #print(nominal)
                             #   debug = iterate_fair.iloc[index, 0]
                            #    if "94" in str(debug):

                    #================Above block of code cleans up nominal======================================



                    #==============Below block of code for cleaning up result==================================
                                multiple_results = False

                                                     #removes things like "6X" at beginning of result

                              #  debug = iterate_fair.iloc[index, 0]  # for debugging
                             #   if "88" in str(debug):

                                if len(result) > 3:
                                    if "X" in str(result[1].upper()):     #removes things like "6X" from front of dim
                                        result_raw = str(result[2:])
                                        result = result_raw.strip(" ")
                                    else:
                                        if "X" in str(result[2].upper()):     #removes things like "6X" from front of dim
                                            result_raw = str(result[3:])
                                            result = result_raw.strip(" ")

                                #print(result)


                                iterate_res = len(result)  # removes end of the result if something like .050x100 (no spaces)
                                while k < iterate_res:
                                    if "X" in str(result[k].upper()):
                                        result = result[:k]
                                        k += iterate_res
                                    k += 1
                                #print(result)


                                k = 0  # removes any symbols, such as ° or "R", except "." from nominal if no space inbetween
                                iterate_res = len(result)
                                while k < iterate_res:
                                    if "." not in str(result[k]) and "-" not in str(result[k]) and "/" not in str(result[k]):
                                        a = result[k]
                                        b = a.isdigit()
                                        if b != True:
                                            counter_letter = result.count(a)
                                            result = result.replace(a, '')
                                            while counter_letter > 0:
                                                iterate_res -= 1
                                                counter_letter -= 1
                                            k -= 1
                                    k += 1
                                #print(result)


                                k = 0                 #get a ranged result
                                iterate_res = len(result)

                                while k < iterate_res:
                                    if "-" in str(result[k]) or "/" in str(result[k]) or "," in str(result[k]):
                                        result_1 = result[0:k].strip(" ")
                                        result_2 = result[k+1:].strip(" ")
                                        multiple_results = True
                                        k += len(result)

                                    k += 1
                                #print(result)
                             #   if "94" in str(debug):
                             #       print(result.strip(" "))
                              #  print(debug)

                    #===================Above block of code for cleaning up result=======================================


                                if ranged_tolerance == True and multiple_results == True:
                                    try:
                                        #for properly doing the math to determine the percentage of tolerance used up
                                        #based on the result, the nominal dimension and the tolerance with a ranged tolerance
                                        #and multiple results for the nominal dimensions, for example if there are 2 results
                                        #because the requirement has "2X".  This will return the result that uses the highest
                                        #percentage of the tolerance only
                                        if float(result_1) > float(nominal):
                                            try:
                                                diff_1 = round((abs(float(nominal) - float(result_1)) / float(upper)) * 100, 2)
                                            except:
                                                pass
                                        else:
                                            if float(result_1) < float(nominal):
                                                try:
                                                    diff_1 = round((abs(float(nominal) - float(result_2)) / float(lower)) * 100, 2)
                                                except:
                                                    pass

                                        if float(result_2) > float(nominal):
                                            try:
                                                diff_2 = round((abs(float(nominal) - float(result_2)) / float(upper)) * 100, 2)
                                            except:
                                                pass
                                        else:
                                            if float(result_2) < float(nominal):
                                                try:
                                                    diff_2 = round((abs(float(nominal) - float(result_1)) / float(lower)) * 100, 2)
                                                except:
                                                    pass

                                        try:
                                            #whichever percentage result is higher append to lists
                                            if diff_1 > diff_2:
                                                percentages.append(diff_1)
                                                item_loc.append(iterate_fair.iloc[index, 0])
                                                ref_loc.append(iterate_fair.iloc[index, 1])
                                                full_requirement.append(requirement)
                                                full_result.append(iterate_fair.iloc[index, 6])
                                                tool_used.append(iterate_fair.iloc[index, 8])
                                            else:
                                                percentages.append(diff_2)
                                                item_loc.append(iterate_fair.iloc[index, 0])
                                                ref_loc.append(iterate_fair.iloc[index, 1])
                                                full_requirement.append(requirement)
                                                full_result.append(iterate_fair.iloc[index, 6])
                                                tool_used.append(iterate_fair.iloc[index, 8])
                                        except:
                                            pass

                                    except:
                                        pass
                                #if ranged tolerance but only 1 result, then perform these calculations/list appends
                                if ranged_tolerance == True and multiple_results == False:
                          #         print(result, nominal, upper, lower)

                                    try:
                                        if float(result) > float(nominal):
                                            try:
                                                diff_1 = round((abs(float(nominal) - float(result)) / float(upper)) * 100, 2)
                                                percentages.append(diff_1)
                                                item_loc.append(iterate_fair.iloc[index, 0])
                                                ref_loc.append(iterate_fair.iloc[index, 1])
                                                full_requirement.append(requirement)
                                                full_result.append(iterate_fair.iloc[index, 6])
                                                tool_used.append(iterate_fair.iloc[index, 8])
                                            except:
                                                pass
                                        else:

                                            if float(result) < float(nominal):
                                                try:
                                                    diff_1 = round((abs(float(nominal) - float(result)) / float(lower)) * 100, 2)
                                                    percentages.append(diff_1)
                                                    item_loc.append(iterate_fair.iloc[index, 0])
                                                    ref_loc.append(iterate_fair.iloc[index, 1])
                                                    full_requirement.append(requirement)
                                                    full_result.append(iterate_fair.iloc[index, 6])
                                                    tool_used.append(iterate_fair.iloc[index, 8])
                                                except:
                                                    pass

                                        if float(result) == float(nominal):
                                            percentages.append(0)
                                            item_loc.append(iterate_fair.iloc[index, 0])
                                            ref_loc.append(iterate_fair.iloc[index, 1])
                                            full_requirement.append(requirement)
                                            full_result.append(iterate_fair.iloc[index, 6])
                                            tool_used.append(iterate_fair.iloc[index, 8])

                                    except:
                                        pass

                                #if not a ranged tolerance
                                if ranged_tolerance == False:
                                    try:
                                        #multiple results for the nominal, then perform this calculation
                                        if multiple_results == True:
                                            #print(nominal, result_1, result_2, tolerance_real)
                                            diff_1 = round((abs(float(nominal) - float(result_1)) / float(tolerance_real)) * 100, 2)
                                            diff_2 = round((abs(float(nominal) - float(result_2)) / float(tolerance_real)) * 100, 2)
                                            #whichever uses a higher percentage of tolerance append to list
                                            if diff_1 > diff_2:
                                                percentages.append(diff_1)
                                            else:
                                                percentages.append(diff_2)

                                        #if not ranged tolerance and only 1 result then perform this calculation
                                        else:
                                            diff = round((abs(float(nominal) - float(result)) / float(tolerance_real)) * 100, 2)
                                            percentages.append(diff)

                                       # if "94" in str(debug):
                                        #    print(nominal, result, tolerance_real)

                                        #append to lists
                                        item_loc.append(iterate_fair.iloc[index, 0])
                                        ref_loc.append(iterate_fair.iloc[index, 1])
                                        full_requirement.append(requirement)
                                        full_result.append(iterate_fair.iloc[index, 6])
                                        tool_used.append(iterate_fair.iloc[index, 8])

                                    except:
                                        pass

            #close dataframe read of excel, this is needed so that I can open the original file and make
            #changes to it .... if not closed properly like this wierd issues happen
            self.dataframe1.close()

    #=================Code Below for reading excel and trasnferring results to FAI and saving FAI to temp folder

            self.updateChanged.emit("Transferring info to temp excel.......")

            #read excel using openpyxl instead of pandas.... not sure you can actually write data to excel files
            #using pandas, so thats why im using openpyxl load_workbook
            wb = load_workbook(self.excel_path, data_only = True)
            all_sheets = wb.sheetnames
            fair3_pages = [x for x in all_sheets if "FAIR 3" in x.upper()]

          #  redFill = PatternFill(start_color='FFEE1111', end_color='FFEE1111', fill_type='solid')

            t = 0
            #read all fair3 sheet in excel and find the correct cell and add data to correct cell
            while t < len(fair3_pages):
                ws = wb[fair3_pages[t]]
                for cell in ws['A']:
                    if cell.value is not None:
                        k = 0
                        while k < len(item_loc):
                            if str(item_loc[k]) == str(cell.value):
                                c1 = ws.cell(row = cell.row, column = 13)
                                c1.value = str(percentages[k]) + "%"
                                if float(percentages[k] > 70):
                                    ws.cell(cell.row, 13).fill = PatternFill(start_color="FFEE1111", fill_type = "solid")
                            k += 1
                t += 1

            self.updateChanged.emit("Saving results to Folder on Desktop.....")

            get_excel_name = self.excel_path.split('\\')
            part_number = get_excel_name[-1]
            part = part_number[0:12]
            ext = part_number[-5:]

            #save workbook and close it
            save_path = TEMP_EXCEL + "\\" + part + " Overlaid Results" + ext
            wb.save(save_path)
            wb.close()

            #if box checkmarked in UI to open excel, this will automatically open it
            if int(self.c) == 2:
                os.system(f'start EXCEL.EXE "{save_path}"')

    #=================Above code for adding the percentage results to FAI and outputting to folder on desktop===============


    #==========================Below code for making a list for over 70% results======================================
            #setup some lists if the percentages are over 70% as this is the data i want to add to the excel

            #this code is just for making a new excel and outputting the results of percentages over 70%
            #so that i can get all the items in 1 easy spot rather than spread out on the original excel file
            over_item_loc = []
            over_ref_loc = []
            over_full_requirement = []
            over_full_result = []
            over_percentages = []
            over_tool_used = []

            k = 0
            while k < len(percentages):
                #if percentage over 70 append data to new lists from the main lists
                if float(percentages[k]) > 70:
                    over_item_loc.append(item_loc[k])
                    over_ref_loc.append(ref_loc[k])
                    over_full_requirement.append(full_requirement[k])
                    over_full_result.append(full_result[k])
                    over_percentages.append(percentages[k])
                    over_tool_used.append(tool_used[k])
                k += 1


            filepath = TEMP_EXCEL + "\\" + part + " over 70% results list" + ext
            wb = openpyxl.Workbook()
            all_sheets = wb.sheetnames
            ws = wb[all_sheets[0]]
            total_results = len(over_item_loc)
            k = 0

            while k < total_results:
                ws.cell(row=k+1, column=1).value = str(over_item_loc[k])
                ws.cell(row=k+1, column=2).value = str(over_ref_loc[k])
                ws.cell(row=k+1, column=3).value = str(over_full_requirement[k])
                ws.cell(row=k+1, column=4).value = str(over_full_result[k])
                ws.cell(row=k+1, column=5).value = str(over_tool_used[k])
                k +=1

            wb.save(filepath)
            wb.close()

            # if box checkmarked in UI to open excel, this will automatically open it
            if int(self.d) == 2:
                os.system(f'start EXCEL.EXE "{filepath}"')

            self.updateChanged.emit("Complete!")
        except:
            easygui.msgbox(msg="ERROR: DUNDUN", title="ERROR")
#===================Above code for making list for over 50% results and saving to excel==========================

#======================Below code for ranged results to get 2 results and tolerance============================\


    #this code is for cleaning up the requirement cell value in the excel and breaking it down to how i want it
    #broken down, this is purely for my needs
    def ranged_req(self, requirement):
        if len(requirement) > 3:
            if "X" in str(requirement[1].upper()):  # removes things like "6X" from front of dim
                requirement = requirement[2:].strip(" ")
            if "X" in str(requirement[2].upper()):  # removes t
                requirement = requirement[3:].strip(" ")

        k = 0
        iterate_nom = len(requirement)  # removes end of the result if something like .050x100 (no spaces)
        while k < iterate_nom:
            if "X" in str(requirement[k].upper()):
                requirement = requirement[:k]
                k += iterate_nom
            k += 1

    #    debug = iterate_fair.iloc[index, 0]
    #    if "44" in str(debug):
    #        print(requirement)

        k = 0  # removes any symbols, such as ° or "R", except "." and "-" and "/" from nominal if no space inbetween
        iterate_nom = len(requirement)
        while k < iterate_nom:
            if "." not in str(requirement[k]) and "/" not in str(requirement[k]):
                a = requirement[k]
                b = a.isdigit()
                if b != True:
                  # print(str(k) + " " + requirement[k])
                    counter_letter = requirement.count(a)
                    requirement = requirement.replace(a, '')
                    while counter_letter > 0:
                        iterate_nom -= 1
                        counter_letter -= 1
                    k -= 1
           # print(requirement)
            k += 1

        nominals_found = False
        k = 0
        iterate_nom = len(requirement)
        while k < iterate_nom:
            if "/" in str(requirement[k]) or "-" in str(requirement[k]):
                nominal_1 = str(requirement[0:k]).strip(" ")
                nominal_2 = str(requirement[k+1:]).strip(" ")
                nominals_found = True
                k += iterate_nom
            k += 1

        if nominals_found == True:
            try:
                #if using the above blocks of code don't output proper results so that the calculation below can be done
                #then return error error that it couldn't be evaluated
                tolerance = abs(round((float(nominal_1) - float(nominal_2)) / 2, 4))

                if float(nominal_1) > float(nominal_2):
                    true_nominal = float(nominal_2) + float(tolerance)
                else:
                    true_nominal = float(nominal_1) + float(tolerance)
                return(true_nominal, tolerance)
            except:
                return("ERROR", "ERROR")
        else:
            return("ERROR", "ERROR")


    #breaking down the cell value so upper and lower ranges of the nominal based on the tolerance are determined
    #for correct calculations when it's a ranged tolerance
    def upper_lower_tol(self, tolerance_raw):
      #  print(tolerance_raw)
        if " " in str(tolerance_raw):  # removes anything after tolerance such as "X 100°"
            k = 0
            end_whileloop = False
            while k < len(tolerance_raw):
                if " " in str(tolerance_raw[k]):
                    tolerance_real = tolerance_raw[0:k]
                    up_k = len(tolerance_raw) - k
                    end_whileloop = True
                if end_whileloop == True:
                    k += up_k
                k += 1
        else:
            tolerance_real = tolerance_raw

      #  debug = iterate_fair.iloc[index, 0]
      #  if "22" in str(debug):
       #     print(tolerance_real)

        k = 0  # if .010x100 no spaces, extracts tolerance at in front of x
        while k < len(tolerance_real):
            if "X" in str(tolerance_real[k]).upper():
                tolerance_real = tolerance_real[0:k]
                k += len(tolerance_real)
            k += 1

     #   if "22" in str(debug):
      #      print(tolerance_real)

        k = 0  # removes any symbols, such as °, except "." from tolerance
        iterate_tol = len(tolerance_real)
        while k < iterate_tol:
            if "." not in str(tolerance_real[k]) and "/" not in str(tolerance_real[k]):
                a = tolerance_real[k]
                b = a.isdigit()
                if b != True:
                    counter_letter = tolerance_real.count(a)
                    tolerance_real = tolerance_real.replace(a, '')
                    while counter_letter > 0:
                        iterate_tol -= 1
                        counter_letter -= 1
                    k -= 1
            k += 1

        upper_low_confirmed = False
        k = 0
        iterate_tol = len(tolerance_real)
        while k < iterate_tol:
            if "/" in str(tolerance_real[k]):
                upper = tolerance_real[0:k].strip(" ")
                lower = tolerance_real[k+1:].strip(" ")
                upper_low_confirmed = True
                k += iterate_tol
            k += 1

        if upper_low_confirmed == True:
            return(upper, lower)
        else:
            return("ERROR", "ERROR")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = Actions()
    sys.exit(app.exec_())